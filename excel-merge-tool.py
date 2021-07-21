import tkinter, tkinter.filedialog, threading
from tkinter.constants import SE
import openpyxl as xl
from copy import copy, deepcopy
from typing import List
from abc import ABC, abstractmethod
from enum import Enum, auto


class DifferentNumberOfColumnsError(Exception):
    pass


class Workbook:

    def __init__(self, workbook_path: str = None):
        if workbook_path is not None:
            self.load_workbook(workbook_path)


    def load_workbook(self, workbook_path: str):
        self.workbook = xl.load_workbook(workbook_path)


    def save_workbook(self, save_path):
        self.workbook.save(save_path)
    

    def get_internal_workbook(self):
        return self.workbook


    @staticmethod
    def is_cell_empty(cell):
        value = cell.value
        return value is None or str(value).strip() == ""
    
    
    @staticmethod
    def first_empty_row(worksheet) -> int:
        last = worksheet.max_row + 10
        for row in range(1, last):
            if Workbook.is_cell_empty(worksheet.cell(row=row, column=1)):
                return row


    @staticmethod
    def get_number_of_rows(worksheet) -> int:
        return Workbook.first_empty_row(worksheet) - 1


    @staticmethod
    def first_empty_column(worksheet) -> int:
        last = worksheet.max_column + 10
        for column in range(1, last):
            if Workbook.is_cell_empty(worksheet.cell(row=1, column=column)):
                return column


    @staticmethod
    def get_number_of_columns(worksheet) -> int:
        return Workbook.first_empty_column(worksheet) - 1


    @staticmethod
    def is_worksheet_empty(worksheet) -> bool:
        return Workbook.is_cell_empty(worksheet.cell(row=1, column=1))


    @staticmethod
    def copy_cell_style(cell_from, cell_to):
        cell_to.font = copy(cell_from.font)
        cell_to.border = copy(cell_from.border)
        cell_to.fill = copy(cell_from.fill)
        cell_to.number_format = copy(cell_from.number_format)
        cell_to.protection = copy(cell_from.protection)
        cell_to.alignment = copy(cell_from.alignment)


    @staticmethod
    def merge_sheets(worksheet1, worksheet2, progress_callback_fn = None, count_rows = True, copy_style = True):
        """ progress_callback_fn takes as input dict:
            {
                "row": {
                    "current": 1,
                    "all": 100
                },
                "column": {
                    "current": 1,
                    "all": 10
                }
            }
        """

        if Workbook.is_worksheet_empty(worksheet1):
            worksheet1 = worksheet2
            return

        if Workbook.is_worksheet_empty(worksheet2):
            return

        column_amount = Workbook.get_number_of_columns(worksheet2)

        if column_amount != Workbook.get_number_of_columns(worksheet1):
            raise DifferentNumberOfColumnsError("Worksheets have different number of columns!")

        worksheet1_row_first_empty = Workbook.first_empty_row(worksheet1)
        
        worksheet2_rows_amount = None
        if count_rows:
            worksheet2_rows_amount = Workbook.get_number_of_rows(worksheet2)
        
        worksheet1_row = worksheet1_row_first_empty
        worksheet2_row = 2
        

        while not Workbook.is_cell_empty(worksheet2.cell(row=worksheet2_row, column=1)):

            for column in range(1, column_amount+1):

                if progress_callback_fn is not None:
                    progress_callback_fn({
                        "row": {
                            "current": worksheet2_row,
                            "all": worksheet2_rows_amount if worksheet2_rows_amount is not None else "?"
                        },
                        "column": {
                            "current": column,
                            "all": column_amount
                        }
                    })

                cell = worksheet1.cell(row=worksheet1_row, column=column)
                cell.value = worksheet2.cell(row=worksheet2_row, column=column).value
                
                if copy_style:
                    if worksheet1_row < 5:
                        style_cell = worksheet2.cell(row=worksheet2_row, column=column)
                    else:
                        style_cell = worksheet1.cell(row=worksheet1_row-2, column=column)
                        
                    Workbook.copy_cell_style(style_cell, cell)
            
            worksheet1_row += 1
            worksheet2_row += 1


    @staticmethod
    def merge_workbook(workbook1, workbook2, progress_callback_fn = None):
        """ progress_callback_fn takes as input dict 
            {
                "sheet": {
                    "current": 0,
                    "all": 10
                },
                "row": {
                    "current": 1,
                    "all": 100
                },
                "column": {
                    "current": 1,
                    "all": 10
                }
            }
        """

        def progress_callback(progress):
            progress["sheet"] = {
                "current": sheet_num + 1,
                "all": len(workbook2.worksheets)
            }
            progress_callback_fn(progress)


        workbook1 = workbook1.get_internal_workbook()
        workbook2 = workbook2.get_internal_workbook()

        for sheet_num, worksheet2 in enumerate(workbook2.worksheets):
            
            is_other_worksheet_in_current_worksheets = worksheet2.title.lower() in map(lambda s: s.lower(), workbook1.sheetnames)
            
            if not is_other_worksheet_in_current_worksheets:
                workbook1.create_sheet(worksheet2.title, workbook2.sheetnames.index(worksheet2.title))
                workbook1[worksheet2.title] = worksheet2
                continue
            
            worksheet1 = workbook1[worksheet2.title]
            Workbook.merge_sheets(worksheet1, worksheet2, progress_callback_fn=progress_callback)


class BtnStatus(Enum):
    DISABLED = "disabled"
    NORMAL = "normal"


class Model:
    files = []
    output_workbook = None
    path_to_save = None
    in_filetypes = (    
                        ("Excel Workbook", "*.xlsx"),
                        ("Excel Macro-Enabled Workbook", "*.xlsm"),
                        ("Excel Macro-Enabled Workbook Template", "*.xltm"),
                        ("Excel Spreadsheet Template", "*.xltx"),
                        # ("OpenDocument Spreadsheet", "*.ods"),
                        # ("CSV (Comma delimited)", "*.csv"),
                        # ("Excel 97-2003 Workbook", "*.xls"),
                        # ("Excel Binary Workbook", "*.xlsb")
                    )

    out_filetypes = (   
                        ("Excel Workbook", "*.xlsx"),
                        # ("OpenDocument Spreadsheet", "*.ods"),
                        # ("CSV (Comma delimited)", "*.csv"),
                    )

class View(ABC):

    @abstractmethod
    def __init__(self, title: str = "Excel Merge Tool", icon_path: str = "./icon.ico"):
        """Initialize window, set up title and icon."""


    @abstractmethod
    def setUpView(self, controller):
        """Set up gui objects."""


    @abstractmethod
    def startMainLoop(self):
        """Start the main loop of the program."""


    @abstractmethod
    def askForInputFiles(self, title: str = "Select workbooks") -> List[str]:
        """Ask user to select input files."""


    @abstractmethod    
    def askForOutputFiles(self, title: str = "Save as") -> List[str]:
        """Ask user to choose where to save output file."""


    @abstractmethod
    def checkBtnMergeStatus(self):
        """Check and set the status of the merge button."""
    

    @abstractmethod
    def setBtnMergeStatus(self, status: BtnStatus):
        """Set the status of the merge button."""


    @abstractmethod
    def updateFileList(self):
        """Update list with input files."""


    @abstractmethod
    def setProgressText(self, text: str):
        """Set the text displayed as the progress."""


    @abstractmethod
    def setSaveLocationText(self, text: str):
        """Set the text with the choosen location of output file."""
    

    @abstractmethod
    def notifySound(self):
        """Play the notification sound."""


class Controller:

    def __init__(self, model: Model, view: View):
        self.model = model
        self.view = view


    def handle_load_files_click(self):
        self.model.files = self.view.askForInputFiles()
        self.model.files = [f.name for f in self.model.files]
        self.view.updateFileList()
        self.view.checkBtnMergeStatus()
    

    def handle_choose_save_location(self):
        self.model.path_to_save = self.view.askForOutputFiles()
        self.view.setSaveLocationText(self.model.path_to_save)
        self.view.checkBtnMergeStatus()


    def handle_merge_click(self):        
        self.view.setBtnMergeStatus(BtnStatus.DISABLED)

        self.view.setProgressText("Loading first workbook...")
        self.model.output_workbook = Workbook(self.model.files[0])

        all_workbooks = len(self.model.files)
        def notify_about_progress(progress):
            text = f'Workbook: {current_workbook+2}/{all_workbooks} | Worksheet: {progress["sheet"]["current"]}/{progress["sheet"]["all"]} | Row: {progress["row"]["current"]}/{progress["row"]["all"]} | Column: {progress["column"]["current"]}/{progress["column"]["all"]}'
            self.view.setProgressText(text)


        # Copy data from each file to out excel file
        for current_workbook, workbook in enumerate(self.model.files[1:]):
            self.view.setProgressText(f"Loading workbook nr {current_workbook+2}...")
            Workbook.merge_workbook(self.model.output_workbook, Workbook(workbook), progress_callback_fn=notify_about_progress)

        # save new excel file
        self.view.setProgressText("Saving now...")
        self.model.output_workbook.save_workbook(save_path=self.model.path_to_save)
        self.view.setProgressText("Done!")
        self.view.notifySound()

        self.view.checkBtnMergeStatus()
    

    def start(self):
        self.view.setUpView(self)
        self.view.startMainLoop()


class View_TK(View):
    # model: Model
    # window: tkinter.Tk
    # btn_merge: tkinter.Button
    # progress_text: tkinter.StringVar
    # save_location_text = tkinter.StringVar


    def __init__(self, model: Model, title: str = "Excel Merge Tool", icon_path: str = "./icon.ico"):
        self.model = model
        self.window = tkinter.Tk()

        self.window.title(title)
        self.window.iconbitmap(icon_path)


    def setUpView(self, controller: Controller):
        btn_load = tkinter.Button(self.window, text="Select files", command=controller.handle_load_files_click, padx=15, pady=5)
        btn_load.grid(row=0,column=0, padx=10, pady=10)

        btn_save = tkinter.Button(self.window, text="Select save location", command=controller.handle_choose_save_location, padx=15, pady=5)
        btn_save.grid(row=0,column=1, padx=10, pady=10)

        self.btn_merge = tkinter.Button(self.window, text="Merge and save", command=lambda : threading.Thread(target=controller.handle_merge_click).start(), padx=15, pady=5, state='disabled')
        self.btn_merge.grid(row=0,column=2, padx=10, pady=10)

        label1 = tkinter.Label(self.window, text="Selected files:")
        label1.grid(row=1, column=0)

        self.list_files_box = tkinter.Listbox(self.window, width=68)
        self.list_files_box.grid(row=2,column=0, columnspan=3, sticky="W", padx=10)

        label2 = tkinter.Label(self.window, text="Save location : ")
        label2.grid(row=3, column=0, pady=5)

        self.save_location_text = tkinter.StringVar(self.window)
        label3 = tkinter.Label(self.window, textvariable=self.save_location_text)
        label3.grid(row=3, column=1, sticky="W", pady=5, columnspan=2)
        self.save_location_text.set("None")

        self.progress_text = tkinter.StringVar(self.window)
        progress_label = tkinter.Label(self.window, textvariable=self.progress_text, padx=10, pady=5)
        progress_label.grid(row=4,column=0, columnspan=3, sticky='W')


    def askForInputFiles(self, title: str = "Select workbooks") -> List[str]:
        return tkinter.filedialog.askopenfiles(title=title, filetypes=self.model.in_filetypes)


    def askForOutputFiles(self, title: str = "Save as") -> List[str]:
        return tkinter.filedialog.asksaveasfilename(title=title, filetypes=self.model.out_filetypes, defaultextension='.xlsx')
    

    def setBtnMergeStatus(self, status: BtnStatus):
        self.btn_merge["state"] = status.value


    def checkBtnMergeStatus(self):
        if self.model.path_to_save is not None and self.model.files is not []:
            self.setBtnMergeStatus(BtnStatus.NORMAL)


    def updateFileList(self):
        self.list_files_box.delete(0, tkinter.END)
        [self.list_files_box.insert(tkinter.END, file) for file in self.model.files]


    def setProgressText(self, text: str):
        self.progress_text.set(text)
    

    def setSaveLocationText(self, text: str):
        self.save_location_text.set(text)


    def notifySound(self):
        self.window.bell()


    def startMainLoop(self):
        self.window.mainloop()


def main():
    model = Model()
    view = View_TK(model)
    controller = Controller(model, view)
    controller.start()


if __name__ == "__main__":
    main()

