import openpyxl as xl
from copy import copy


class DifferentNumberOfColumnsError(Exception):
    pass


class Workbook:

    def __init__(self, workbook_path: str = None):
        if workbook_path is not None:
            self.load_workbook(workbook_path)


    def load_workbook(self, workbook_path: str):
        self.workbook = xl.load_workbook(workbook_path)


    def save_workbook(self, save_path):
        self.workbook.save(save_path)
    

    def get_internal_workbook(self):
        return self.workbook


    @staticmethod
    def is_cell_empty(cell):
        value = cell.value
        return value is None or str(value).strip() == ""
    
    
    @staticmethod
    def first_empty_row(worksheet) -> int:
        last = worksheet.max_row + 10
        for row in range(1, last):
            if Workbook.is_cell_empty(worksheet.cell(row=row, column=1)):
                return row


    @staticmethod
    def get_number_of_rows(worksheet) -> int:
        return Workbook.first_empty_row(worksheet) - 1


    @staticmethod
    def first_empty_column(worksheet) -> int:
        last = worksheet.max_column + 10
        for column in range(1, last):
            if Workbook.is_cell_empty(worksheet.cell(row=1, column=column)):
                return column


    @staticmethod
    def get_number_of_columns(worksheet) -> int:
        return Workbook.first_empty_column(worksheet) - 1


    @staticmethod
    def is_worksheet_empty(worksheet) -> bool:
        return Workbook.is_cell_empty(worksheet.cell(row=1, column=1))


    @staticmethod
    def copy_cell_style(cell_from, cell_to):
        cell_to.font = copy(cell_from.font)
        cell_to.border = copy(cell_from.border)
        cell_to.fill = copy(cell_from.fill)
        cell_to.number_format = copy(cell_from.number_format)
        cell_to.protection = copy(cell_from.protection)
        cell_to.alignment = copy(cell_from.alignment)


    @staticmethod
    def merge_sheets(worksheet1, worksheet2, progress_callback_fn = None, count_rows = True, copy_style = True):
        """ progress_callback_fn takes as input dict:
            {
                "row": {
                    "current": 1,
                    "all": 100
                },
                "column": {
                    "current": 1,
                    "all": 10
                }
            }
        """

        if Workbook.is_worksheet_empty(worksheet1):
            worksheet1 = worksheet2
            return

        if Workbook.is_worksheet_empty(worksheet2):
            return

        column_amount = Workbook.get_number_of_columns(worksheet2)

        if column_amount != Workbook.get_number_of_columns(worksheet1):
            raise DifferentNumberOfColumnsError("Worksheets have different number of columns!")

        worksheet1_row_first_empty = Workbook.first_empty_row(worksheet1)
        
        worksheet2_rows_amount = None
        if count_rows:
            worksheet2_rows_amount = Workbook.get_number_of_rows(worksheet2)
        
        worksheet1_row = worksheet1_row_first_empty
        worksheet2_row = 2
        

        while not Workbook.is_cell_empty(worksheet2.cell(row=worksheet2_row, column=1)):

            for column in range(1, column_amount+1):

                if progress_callback_fn is not None:
                    progress_callback_fn({
                        "row": {
                            "current": worksheet2_row,
                            "all": worksheet2_rows_amount if worksheet2_rows_amount is not None else "?"
                        },
                        "column": {
                            "current": column,
                            "all": column_amount
                        }
                    })

                cell = worksheet1.cell(row=worksheet1_row, column=column)
                cell.value = worksheet2.cell(row=worksheet2_row, column=column).value
                
                if copy_style:
                    if worksheet1_row < 5:
                        style_cell = worksheet2.cell(row=worksheet2_row, column=column)
                    else:
                        style_cell = worksheet1.cell(row=worksheet1_row-2, column=column)
                        
                    Workbook.copy_cell_style(style_cell, cell)
            
            worksheet1_row += 1
            worksheet2_row += 1


    @staticmethod
    def merge_workbook(workbook1, workbook2, progress_callback_fn = None):
        """ progress_callback_fn takes as input dict 
            {
                "sheet": {
                    "current": 0,
                    "all": 10
                },
                "row": {
                    "current": 1,
                    "all": 100
                },
                "column": {
                    "current": 1,
                    "all": 10
                }
            }
        """

        def progress_callback(progress):
            progress["sheet"] = {
                "current": sheet_num + 1,
                "all": len(workbook2.worksheets)
            }
            progress_callback_fn(progress)


        workbook1 = workbook1.get_internal_workbook()
        workbook2 = workbook2.get_internal_workbook()

        for sheet_num, worksheet2 in enumerate(workbook2.worksheets):
            
            is_other_worksheet_in_current_worksheets = worksheet2.title.lower() in map(lambda s: s.lower(), workbook1.sheetnames)
            
            if not is_other_worksheet_in_current_worksheets:
                workbook1.create_sheet(worksheet2.title, workbook2.sheetnames.index(worksheet2.title))
                workbook1[worksheet2.title] = worksheet2
                continue
            
            worksheet1 = workbook1[worksheet2.title]
            Workbook.merge_sheets(worksheet1, worksheet2, progress_callback_fn=progress_callback)
